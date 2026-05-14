"""
main.py — Furnace Flame Temperature Mapping API
================================================

Encryption
----------
  Layer 2 ONLY — At-Rest AES-256-CBC + HMAC-SHA256
  Data is encrypted when written to the DB and decrypted when read.
  Transit is plain HTTP — no WebCrypto required on the client.
"""



from fastapi import FastAPI, UploadFile, File, Form, Query
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import pymysql
import json
import uuid
import shutil
import os
from datetime import datetime
import re
from pydantic import BaseModel, Field
from typing import Optional, Any, Dict, List

from dateutil import parser as date_parser

from crypto import (
    encrypt_at_rest,
    decrypt_at_rest_float,
)

app = FastAPI()

# ─── CORS ─────────────────────────────────────────────────────
origins = [
    "http://localhost:5173",
    "https://furnaceflametemperaturemapping.vercel.app",
    "http://101.53.132.91:5173",
    "http://91.203.132.34:5173",
    "https://91.203.132.34:5174",
    "https://localhost:5174",
    "http://91.203.132.34:5176",
    "http://localhost:5176",

]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── MYSQL CONNECTION ──────────────────────────────────────────
def get_db():
    import mysql.connector
    return mysql.connector.connect(
        host='localhost', user='root', password='Vishalgad5@3332', database='furnace_db'
    )


UPLOAD_DIR = "uploads"
os.makedirs(UPLOAD_DIR, exist_ok=True)


def clean(val):
    """Return stripped string or None for blank/NaN."""
    if val is None:
        return None
    if isinstance(val, float) and pd.isna(val):
        return None
    s = str(val).strip()
    return s if s else None

def parse_date_flexible(value: Any) -> Optional[datetime.date]:
    """
    Parse a date from many possible inputs (Excel cells, strings, etc.)
    and normalize to a `date` object.
    """
    if value is None:
        return None

    # Common Excel / pandas cases
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        # includes `datetime.date` and some pandas date-like values
        try:
            return datetime(int(value.year), int(value.month), int(value.day)).date()
        except Exception:
            pass

    s = clean(value)
    if not s:
        return None

    # Strip obvious noise and normalize separators a bit
    s = re.sub(r"\s+", " ", s)

    try:
        # dayfirst=True keeps existing DD/MM/YYYY behavior for ambiguous cases,
        # but still allows many formats (e.g. "May 8 2026", "2026-05-08", etc.)
        dt = date_parser.parse(s, dayfirst=True, fuzzy=True)
        return dt.date()
    except Exception:
        return None

def extract_date_from_sheet(df):
    """
    Look for a cell matching 'Date :- DD/MM/YYYY' in the first 5 rows.
    Returns a date object or None.
    """
    def _looks_like_date_text(s: str) -> bool:
        s = s.strip()
        if not s:
            return False
        # 4-digit year, or dd/mm/yy-style with separators, or month names
        if re.search(r"\b(19|20)\d{2}\b", s):
            return True
        if re.search(r"\b\d{1,2}\s*[/\-.]\s*\d{1,2}\s*[/\-.]\s*\d{2,4}\b", s):
            return True
        if re.search(r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)\b", s, re.IGNORECASE):
            return True
        return False

    # First pass: find a "Date" label and parse from the same cell or the cell to the right
    for i in range(min(5, len(df))):
        row_vals = list(df.iloc[i])
        for j, val in enumerate(row_vals):
            if val is None:
                continue
            text = str(val)
            if re.search(r"\bdate\b", text, re.IGNORECASE):
                # Sometimes date is embedded in the same cell: "Date :- 14.05.2026"
                parsed = parse_date_flexible(text)
                if parsed:
                    return parsed
                # Common template pattern: "Date :-" | "14.05.2026"
                if j + 1 < len(row_vals):
                    parsed_next = parse_date_flexible(row_vals[j + 1])
                    if parsed_next:
                        return parsed_next

    # Second pass: any obvious-looking date in the first 5 rows
    for i in range(min(5, len(df))):
        for val in df.iloc[i]:
            if val is None:
                continue
            parsed = parse_date_flexible(val)
            if parsed:
                # Guard against accidentally parsing non-date numeric fields:
                # only accept stringy values that resemble a date.
                if isinstance(val, str):
                    if _looks_like_date_text(val):
                        return parsed
                else:
                    return parsed

    return None


def find_row(df, text):
    """
    Return the integer iloc index of the first row where any cell
    contains `text` (case-insensitive). Returns None if not found.
    """
    text_l = text.lower()
    for i, row in df.iterrows():
        for cell in row:
            if cell is not None and text_l in str(cell).lower():
                return i
    return None

def find_col_in_row(df, row_idx, text):
    """
    Return the column position (0-based) in df.iloc[row_idx] where the
    cell contains `text` (case-insensitive). Returns None if not found.
    """
    text_l = text.lower()
    for col_pos, val in enumerate(df.iloc[row_idx]):
        if val is not None and text_l in str(val).lower():
            return col_pos
    return None

def extract_metadata(df):
    meta = {}

    for i in range(min(6, len(df))):
        row_vals = list(df.iloc[i])
        row_str  = ' '.join(str(v) for v in row_vals if v is not None)

        # BSL + unit are in row 2 (iloc 1)
        if i == 1:
            meta['bsl'] = clean(row_vals[0])
            unit_raw = clean(row_vals[1])
            if unit_raw:
                um = re.search(r'\d+', unit_raw)
                meta['unit'] = um.group() if um else unit_raw

        # ── Date: find label cell, read value from the NEXT cell ──────────
        if 'run_date' not in meta:
            for col_pos, val in enumerate(row_vals):
                if val is not None and re.search(r'date\s*[:\-]+', str(val), re.IGNORECASE):
                    # Value is in the very next cell to the right
                    next_val = clean(row_vals[col_pos + 1]) if col_pos + 1 < len(row_vals) else None
                    if next_val:
                        parsed = parse_date_flexible(next_val)
                        if parsed:
                            meta['run_date'] = parsed
                    break  # stop scanning columns once label found

        # ── Time: same pattern ────────────────────────────────────────────
        if 'run_time' not in meta:
            for col_pos, val in enumerate(row_vals):
                if val is not None and re.search(r'time\s*[:\-]+', str(val), re.IGNORECASE):
                    next_val = clean(row_vals[col_pos + 1]) if col_pos + 1 < len(row_vals) else None
                    if next_val and re.match(r'\d{1,2}:\d{2}', next_val):
                        meta['run_time'] = next_val
                    break

        # Numeric KV pairs (unchanged)
        for col_pos, val in enumerate(row_vals):
            if val is None:
                continue
            vs = str(val).strip()
            nv = clean(row_vals[col_pos + 1]) if col_pos + 1 < len(row_vals) else None

            if re.search(r'load\s*\(mw\)', vs, re.IGNORECASE) and nv:
                meta['load_mw'] = nv
            if re.search(r'total coal flow', vs, re.IGNORECASE) and nv:
                meta['coal_flow_tph'] = nv
            if re.search(r'total air flow', vs, re.IGNORECASE) and nv:
                meta['air_flow_tph'] = nv
            if re.search(r'coal mills in service', vs, re.IGNORECASE) and nv:
                meta['coal_mills_in_service'] = nv
            if re.search(r'oil guns in service', vs, re.IGNORECASE) and nv:
                meta['oil_guns_in_service'] = nv
            if re.search(r'burner tilt', vs, re.IGNORECASE) and nv:
                meta['burner_tilt_position'] = nv

    return meta

def extract_boiler_mill_params(df):
    """
    Dynamically find the Boiler & Mill Parameters section, then scan
    both the left block and right block for L/R values by label text.
    """
    boiler_row = find_row(df, 'BOILER & MILL PARAMETERS')
    if boiler_row is None:
        return None
 
    coal_row = find_row(df, 'COAL MILL PARAMETERS')
    end_row  = coal_row if coal_row is not None else len(df)
 
    # ── Detect column positions dynamically ───────────────────────────────
    # Left block: the section header row tells us where 'L' and 'R' are
    # Screenshot: cols are A(0)=label, B(1)=blank, C(2)=L, D(3)=R,
    #             E(4)=right-label, F(5)=blank, G(6)=L, H(7)=R
    # We find 'L' and 'R' header cols in the boiler header row.
 
    header_row = boiler_row  # "Boiler & Mill Parameters :" is the section row
    # Look one row below for the L / R sub-headers
    lr_row = boiler_row + 1
    lr_vals = list(df.iloc[lr_row]) if lr_row < len(df) else []
 
    # Find columns with 'L' and 'R' (there will be two pairs: left block & right block)
    l_cols = [c for c, v in enumerate(lr_vals) if clean(v) == 'L']
    r_cols = [c for c, v in enumerate(lr_vals) if clean(v) == 'R']
 
    # Left block uses first L/R pair, right block uses second pair
    left_label_col  = l_cols[0] - 1 if l_cols else 0
    left_l_col      = l_cols[0]     if l_cols else 2
    left_r_col      = r_cols[0]     if r_cols else 3
    right_label_col = l_cols[1] - 1 if len(l_cols) > 1 else 4
    right_l_col     = l_cols[1]     if len(l_cols) > 1 else 6
    right_r_col     = r_cols[1]     if len(r_cols) > 1 else 7
 
    data_start = lr_row + 1   # actual data starts one row after L/R header
 
    left_data  = _scan_boiler_block(df, data_start, end_row,
                                    left_label_col,  left_l_col,  left_r_col,  _BOILER_LEFT)
    right_data = _scan_boiler_block(df, data_start, end_row,
                                    right_label_col, right_l_col, right_r_col, _BOILER_RIGHT)
 
    return {**left_data, **right_data}

_COAL_ROW_LABELS = {
    # result key : label fragment
    'coal_flow_tph':    'coal flow',
    'pa_flow_tph':      'pa flow',
    'mill_dp_mmwc':     'mill dp',
    'mill_outlet_temp': 'mill outlet',
    'mill_current_amp': 'mill current',
}


def extract_coal_mill_params(df):
    """
    Find Coal Mill Parameters section. Detect mill names from the COAL MILLS
    header row, then for each parameter row scan by label.
 
    Returns list of dicts: {mill, coal_flow_tph, pa_flow_tph, ...}
    """
    coal_row = find_row(df, 'COAL MILL PARAMETERS')
    if coal_row is None:
        return []
 
    # Find end of section
    end_row = find_row(df, 'end')  # Your Excel has 'end' at the bottom
    if end_row is None:
        end_row = len(df)
    
    # The COAL MILLS row has mill names: "Coal Mill - A", "Coal Mill - B", ...
    mills_header_row = None
    for i in range(coal_row, min(coal_row + 5, len(df))):
        row_str = ' '.join(str(v) for v in df.iloc[i] if v is not None).lower()
        if 'coal mill -' in row_str or 'coal mills' in row_str:
            mills_header_row = i
            break
 
    if mills_header_row is None:
        return []
 
    # Collect (mill_letter, col_index) pairs from that header row
    mills = []
    for col_pos, val in enumerate(df.iloc[mills_header_row]):
        if val is None:
            continue
        vs = str(val).strip()
        
        # Improved regex to match "Coal Mill - A" or "Coal Mill A"
        m = re.search(r'coal\s*mill\s*[-–—]?\s*([A-Ha-h])', vs, re.IGNORECASE)
        if m:
            mills.append((m.group(1).upper(), col_pos))
    
    if not mills:
        return []
 
    # Find the label column (where parameter names are)
    label_col = 0  # usually column A (index 0)
    
    # Try to find the label column dynamically
    for col_pos in range(min(3, df.shape[1])):
        sample_val = str(df.iloc[mills_header_row + 1, col_pos]) if mills_header_row + 1 < len(df) else ""
        if 'coal flow' in sample_val.lower():
            label_col = col_pos
            break
 
    # For each row label, find the row index
    label_row_map = {}
    for i in range(mills_header_row + 1, min(end_row, len(df))):
        cell = clean(df.iloc[i, label_col])
        if not cell:
            continue
        cell_l = cell.lower()
        for key, frag in _COAL_ROW_LABELS.items():
            if frag in cell_l and key not in label_row_map:
                label_row_map[key] = i
 
    # Build result – one dict per mill
    result = []
    for mill_letter, col_pos in mills:
        entry = {'mill': mill_letter}
        for key in _COAL_ROW_LABELS:
            row_i = label_row_map.get(key)
            if row_i is not None and col_pos < df.shape[1]:
                value = clean(df.iloc[row_i, col_pos])
                # Try to convert to number
                if value is not None:
                    try:
                        entry[key] = float(value)
                    except (ValueError, TypeError):
                        entry[key] = value
                else:
                    entry[key] = None
            else:
                entry[key] = None
        # Only include if at least one value is non-null
        if any(v for k, v in entry.items() if k != 'mill' and v is not None):
            result.append(entry)
    
    return result

def extract_profile_points(df):
    """
    Find the ELEVATION header row, then collect rows until we hit the
    Boiler & Mill Parameters section (or a row with no elevation value).
    
    Dynamically detects all corner columns (Corner 1, Corner 2, etc.)
    Returns list of dicts: {elevation, c1, c2, c3, c4, ...cN, avg_val}
    """
    elev_row = find_row(df, 'ELEVATION')
    if elev_row is None:
        return []
    
    # Get the header row to detect all column positions
    header_row = df.iloc[elev_row]
    
    # Find ELEVATION column position
    elev_col = find_col_in_row(df, elev_row, 'ELEVATION')
    if elev_col is None:
        elev_col = 0
    
    # Dynamically detect all corner columns
    corner_cols = []  # List of (column_name, column_index)
    avg_col = None
    
    for col_pos in range(elev_col + 1, len(header_row)):
        cell_val = clean(header_row.iloc[col_pos])
        if cell_val is None:
            continue
        
        cell_str = str(cell_val).strip().upper()
        
        # Match corner columns: CORNER 1, CORNER 2, CORNER 3, CORNER 4, CORNER 5, etc.
        corner_match = re.search(r'CORNER\s*(\d+)', cell_str)
        if corner_match:
            corner_num = corner_match.group(1)
            corner_cols.append((f'c{corner_num}', col_pos))
            continue
        
        # Also match just C1, C2, C3, C4, C5 format
        corner_match = re.search(r'^C(\d+)$', cell_str)
        if corner_match:
            corner_num = corner_match.group(1)
            corner_cols.append((f'c{corner_num}', col_pos))
            continue
        
        # Match average column
        if re.search(r'AVG|AVERAGE', cell_str):
            avg_col = col_pos
    
    # If no corner columns detected, fall back to default c1-c4 positions
    if not corner_cols:
        corner_cols = [
            ('c1', elev_col + 1),
            ('c2', elev_col + 2),
            ('c3', elev_col + 3),
            ('c4', elev_col + 4),
        ]
        if avg_col is None:
            avg_col = elev_col + 5
    
    # Find where the boiler section starts so we know when to stop
    boiler_row = find_row(df, 'BOILER & MILL PARAMETERS')
    end_row = boiler_row if boiler_row is not None else len(df)
    
    points = []
    for i in range(elev_row + 1, end_row):
        row = df.iloc[i]
        elev_val = clean(row.iloc[elev_col])
        
        # Stop on NOTE row or blank elevation
        if elev_val is None:
            continue
        if re.search(r'\bNOTE\b', str(elev_val), re.IGNORECASE):
            break
        
        # Try to parse as a number
        try:
            # Extract numeric part from elevation (e.g., "58.4" from "GH/ 35.0")
            elev_num = re.search(r'[\d.]+', str(elev_val))
            if elev_num:
                float(elev_num.group())
            else:
                continue
        except ValueError:
            continue  # skip label-only rows like "GH / 35.0 M"
        
        # Build point dictionary with all corner columns
        point = {
            'elevation': elev_val,
        }
        
        # Add all corner columns
        for col_name, col_idx in corner_cols:
            if col_idx < len(row):
                point[col_name] = clean(row.iloc[col_idx])
            else:
                point[col_name] = None
        
        # Add average
        if avg_col is not None and avg_col < len(row):
            point['avg'] = clean(row.iloc[avg_col])
        else:
            # Calculate average from available corners if avg column not found
            corner_values = []
            for col_name, col_idx in corner_cols:
                val = point.get(col_name)
                if val is not None:
                    try:
                        corner_values.append(float(val))
                    except (ValueError, TypeError):
                        pass
            if corner_values:
                point['avg'] = str(sum(corner_values) / len(corner_values))
            else:
                point['avg'] = None
        
        points.append(point)
    
    return points


_BOILER_LEFT = {
    # key in result dict : label fragment to search for
    'main_steam_pressure_l': ('main steam pressure', 'l'),
    'main_steam_pressure_r': ('main steam pressure', 'r'),
    'main_steam_flow_l':     ('main steam flow',     'l'),
    'main_steam_flow_r':     ('main steam flow',     'r'),
    'superheat_spray_l':     ('superheat spray',     'l'),
    'superheat_spray_r':     ('superheat spray',     'r'),
    'reheat_spray_l':        ('re-heat spray',       'l'),
    'reheat_spray_r':        ('re-heat spray',       'r'),
    'o2_aph_inlet_pcr_l':   ('o2',                  'l'),
    'o2_aph_inlet_pcr_r':   ('o2',                  'r'),
    'wind_box_dp_l':         ('wind box',            'l'),
    'wind_box_dp_r':         ('wind box',            'r'),
    'total_pa_flow':         ('total pa flow',       None),
}

_BOILER_RIGHT = {
    'fg_temp_after_dpsh_l':  ('dpsh', 'l'),
    'fg_temp_after_dpsh_r':  ('dpsh', 'r'),
    'fg_temp_after_psh_l':   ('psh',  'l'),
    'fg_temp_after_psh_r':   ('psh',  'r'),
    'fg_temp_after_rh_l':    ('rh',   'l'),   # NOTE: was swapped in original code
    'fg_temp_after_rh_r':    ('rh',   'r'),
    'fg_temp_after_hsh_l':   ('hsh',  'l'),
    'fg_temp_after_hsh_r':   ('hsh',  'r'),
    'fg_temp_after_eco_l':   ('eco',  'l'),
    'fg_temp_after_eco_r':   ('eco',  'r'),
    'fg_temp_after_aph_l':   ('aph',  'l'),
    'fg_temp_after_aph_r':   ('aph',  'r'),
}
def enc(v):
    return v


def upsert_boiler_mill_params(cur, run_id, params):
    cur.execute("DELETE FROM boiler_mill_params WHERE run_id = %s", (run_id,))
    cur.execute(
        """
        INSERT INTO boiler_mill_params (
            run_id,
            main_steam_pressure_l, main_steam_pressure_r,
            main_steam_flow_l,     main_steam_flow_r,
            superheat_spray_l,     superheat_spray_r,
            reheat_spray_l,        reheat_spray_r,
            o2_aph_inlet_pcr_l,   o2_aph_inlet_pcr_r,
            wind_box_dp_l,         wind_box_dp_r,
            total_pa_flow,
            fg_temp_after_dpsh_l,  fg_temp_after_dpsh_r,
            fg_temp_after_psh_l,   fg_temp_after_psh_r,
            fg_temp_after_rh_l,    fg_temp_after_rh_r,
            fg_temp_after_hsh_l,   fg_temp_after_hsh_r,
            fg_temp_after_eco_l,   fg_temp_after_eco_r,
            fg_temp_after_aph_l,   fg_temp_after_aph_r,
            created_at
        ) VALUES (
            %s,
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s,
            %s, %s, %s, %s, %s, %s, %s, %s,
            %s, %s, %s, %s,
            NOW()
        )
        """,
        (
            run_id,
            params.get('main_steam_pressure_l'), params.get('main_steam_pressure_r'),
            params.get('main_steam_flow_l'),     params.get('main_steam_flow_r'),
            params.get('superheat_spray_l'),     params.get('superheat_spray_r'),
            params.get('reheat_spray_l'),        params.get('reheat_spray_r'),
            params.get('o2_aph_inlet_pcr_l'),    params.get('o2_aph_inlet_pcr_r'),
            params.get('wind_box_dp_l'),         params.get('wind_box_dp_r'),
            params.get('total_pa_flow'),
            params.get('fg_temp_after_dpsh_l'),  params.get('fg_temp_after_dpsh_r'),
            params.get('fg_temp_after_psh_l'),   params.get('fg_temp_after_psh_r'),
            params.get('fg_temp_after_rh_l'),    params.get('fg_temp_after_rh_r'),   # ← fixed order
            params.get('fg_temp_after_hsh_l'),   params.get('fg_temp_after_hsh_r'),
            params.get('fg_temp_after_eco_l'),   params.get('fg_temp_after_eco_r'),
            params.get('fg_temp_after_aph_l'),   params.get('fg_temp_after_aph_r'),
        ),
    )


def upsert_coal_mill_params(cur, run_id, mills):
    cur.execute("DELETE FROM coal_mill_params WHERE run_id = %s", (run_id,))
    for m in mills:
        cur.execute(
            """
            INSERT INTO coal_mill_params (
                run_id, mill,
                coal_flow_tph, pa_flow_tph, mill_dp_mmwc,
                mill_outlet_temp, mill_current_amp,
                created_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, NOW())
            """,
            (
                run_id,
                m.get('mill'),
                m.get('coal_flow_tph'),
                m.get('pa_flow_tph'),
                m.get('mill_dp_mmwc'),
                m.get('mill_outlet_temp'),
                m.get('mill_current_amp'),
            ),
        )

# ── Decrypt helpers ────────────────────────────────────────────
_BOILER_FIELDS = [
    "main_steam_pressure_l", "main_steam_pressure_r",
    "main_steam_flow_l",     "main_steam_flow_r",
    "superheat_spray_l",     "superheat_spray_r",
    "reheat_spray_l",        "reheat_spray_r",
    "o2_aph_inlet_pcr_l",   "o2_aph_inlet_pcr_r",
    "wind_box_dp_l",         "wind_box_dp_r",
    "total_pa_flow",
    "fg_temp_after_dpsh_l",  "fg_temp_after_dpsh_r",
    "fg_temp_after_psh_l",   "fg_temp_after_psh_r",
    "fg_temp_after_rh_l",    "fg_temp_after_rh_r",
    "fg_temp_after_hsh_l",   "fg_temp_after_hsh_r",
    "fg_temp_after_eco_l",   "fg_temp_after_eco_r",
    "fg_temp_after_aph_l",   "fg_temp_after_aph_r",
]
_COAL_FIELDS = ["coal_flow_tph", "pa_flow_tph", "mill_dp_mmwc", "mill_outlet_temp", "mill_current_amp"]


def _dec_boiler(row):
    if not row:
        return row
    out = dict(row)
    for f in _BOILER_FIELDS:
        if f in out:
            out[f] = decrypt_at_rest_float(out[f])
    return out

def _scan_boiler_block(df, section_start, section_end, label_col, l_col, r_col, field_map):
    """
    Scan rows [section_start, section_end) looking for label text in `label_col`.
    For each match, grab values from l_col and r_col.
    field_map: {result_key: (label_fragment, 'l'|'r'|None)}
    Returns a flat dict of {result_key: value}.
    """
    # Build a lookup: label_fragment → list of rows where it appears
    rows_by_label = {}
    for i in range(section_start, section_end):
        cell = clean(df.iloc[i, label_col])
        if not cell:
            continue
        cell_l = cell.lower()
        for result_key, (frag, side) in field_map.items():
            if frag in cell_l:
                rows_by_label.setdefault(frag, []).append(i)
 
    result = {}
    for result_key, (frag, side) in field_map.items():
        rows = rows_by_label.get(frag, [])
        if not rows:
            result[result_key] = None
            continue
        row_i = rows[0]   # take first match
        if side == 'l':
            result[result_key] = clean(df.iloc[row_i, l_col]) if l_col < df.shape[1] else None
        elif side == 'r':
            result[result_key] = clean(df.iloc[row_i, r_col]) if r_col < df.shape[1] else None
        else:
            # No side (e.g. total_pa_flow which is a merged/single value)
            result[result_key] = clean(df.iloc[row_i, l_col]) if l_col < df.shape[1] else None
 
    return result


def _dec_coal(row):
    out = dict(row)
    for f in _COAL_FIELDS:
        if f in out:
            out[f] = decrypt_at_rest_float(out[f])
    return out


def _dec_profile(row):
    """Decrypt all numerical fields in a profile row"""
    if not row:
        return row
    out = dict(row)
    # Decrypt elevation and avg_val
    for f in ["elevation", "avg_val", "avg"]:
        if f in out and out[f] is not None:
            out[f] = decrypt_at_rest_float(out[f])
    
    # Decrypt all corner columns (c1, c2, c3, c4, c5, c6, c7, c8...)
    for key in list(out.keys()):
        if re.match(r'^c\d+$', key):  # Match c1, c2, c3, c4, c5, etc.
            if out[key] is not None:
                out[key] = decrypt_at_rest_float(out[key])
    
    return out


# ─── UPLOAD ───────────────────────────────────────────────────
@app.post('/upload')
async def upload_file(
    file:       UploadFile = File(...),
    station_id: int        = Form(1),
    unit_id:    int        = Form(1),
    uploaded_by: str       = Form('system'),
    notes:      str        = Form(''),
):
    try:
        file_id   = str(uuid.uuid4())
        file_path = os.path.join(UPLOAD_DIR, f'{file_id}_{file.filename}')
        with open(file_path, 'wb') as f:
            shutil.copyfileobj(file.file, f)
 
        # Read all sheets, no header inference
        df_dict = pd.read_excel(file_path, sheet_name=None, header=None)
 
        conn = get_db()
        cur  = conn.cursor(dictionary=True)   # ← dictionary=True so fetchall returns dicts
        results = []
 
        for sheet_name, df in df_dict.items():
            # ── Date ──────────────────────────────────────────────────────
            run_date = extract_date_from_sheet(df)
            if not run_date:
                continue
 
            # ── Metadata ──────────────────────────────────────────────────
            meta     = extract_metadata(df)
            location = meta.get('bsl', '')
            unit     = meta.get('unit', '')
 
            # ── Profile points ────────────────────────────────────────────
            points = extract_profile_points(df)
            if not points:
                continue
 
            # ── Boiler & Mill params ──────────────────────────────────────
            boiler_params = extract_boiler_mill_params(df)
 
            # ── Coal mill params ──────────────────────────────────────────
            coal_mills = extract_coal_mill_params(df)
 
            # ── Persist run ───────────────────────────────────────────────
            cur.callproc('sp_create_run', (
                station_id, unit_id, file.filename,
                datetime.now(), run_date, uploaded_by, notes,
                location, unit,
            ))
            # callproc stores results; fetch from the result set
            run_id = None
            for result_set in cur.stored_results():
                row = result_set.fetchone()
                if row:
                    run_id = row['run_id']
                    break
 
            if run_id is None:
                continue
 
            # ── Persist profile points ────────────────────────────────────
            cur.callproc('sp_add_run_points_bulk', (
                run_id, location, unit_id, json.dumps(points),
            ))
 
            # ── Persist boiler params ─────────────────────────────────────
            if boiler_params:
                upsert_boiler_mill_params(cur, run_id, boiler_params)
 
            # ── Persist coal mill params ──────────────────────────────────
            if coal_mills:
                upsert_coal_mill_params(cur, run_id, coal_mills)
 
            results.append({
                'sheet':          sheet_name,
                'run_id':         run_id,
                'date':           str(run_date),
                'point_rows':     len(points),
                'boiler_params':  boiler_params is not None,
                'coal_mill_rows': len(coal_mills),
            })
 
        conn.commit()
        cur.close()
        conn.close()
 
        return {
            'message':               'Upload processed successfully',
            'total_sheets_processed': len(results),
            'runs':                  results,
        }
 
    except Exception as exc:
        import traceback
        return {'error': str(exc), 'trace': traceback.format_exc()}

# ─── HISTORY ──────────────────────────────────────────────────
@app.get("/history")
def get_history(
    station_id: str,
    unit_id: int,
    start_date: Optional[str] = Query(None),
    end_date: Optional[str] = Query(None),
):
    conn = get_db()
    cur = conn.cursor(dictionary=True)
    
    sd_obj = parse_date_flexible(start_date) if start_date else None
    ed_obj = parse_date_flexible(end_date) if end_date else None
    sd = sd_obj if sd_obj else "2000-01-01"
    ed = ed_obj if ed_obj else "2100-01-01"

    s = (station_id or "").strip()
    rows = []

    if re.fullmatch(r"\d+", s):
        cur.callproc("sp_get_runs", (int(s), unit_id, sd, ed))
        for result_set in cur.stored_results():
            rows = result_set.fetchall()
    else:
        # Use DATE() function to compare only the date part
        cur.execute(
            """
            SELECT *
            FROM runs
            WHERE unit_id = %s
              AND location = %s
              AND DATE(run_date) BETWEEN %s AND %s
            ORDER BY run_date DESC, run_timestamp DESC
            """,
            (unit_id, s, sd, ed),
        )
        rows = cur.fetchall()

        # If still no results, try without date filter to see if data exists
        if not rows:
            print(f"No results with date filter, checking if data exists for {s}, {unit_id}")
            cur.execute(
                """
                SELECT run_id, run_date, location, unit_id
                FROM runs
                WHERE unit_id = %s AND location = %s
                """,
                (unit_id, s)
            )
            existing = cur.fetchall()
            print(f"Found {len(existing)} runs without date filter: {existing}")
    
    cur.close()
    conn.close()
    
    print(f"Returning {len(rows)} rows for {s}, unit {unit_id}, dates {sd} to {ed}")
    
    return rows

    
@app.get("/history/{run_id}")
def get_run(run_id: int):
    try:
        conn = get_db()
        cur = conn.cursor(dictionary=True)
        
        # Call the stored procedure
        cur.callproc("sp_get_run_profile", (run_id,))
        
        # Fetch results - important: stored_results() must be iterated
        rows = []
        for result_set in cur.stored_results():
            rows = result_set.fetchall()
            break  # Take the first result set
        
        cur.close()
        conn.close()
        
        print(f"DEBUG: Found {len(rows)} rows for run_id {run_id}")
        
        if not rows:
            return {
                "elevation": [],
                "average": [],
            }
        
        # Build the response
        result = {
            "elevation": [float(r["elevation"]) for r in rows],
            "average": [float(r["avg_val"]) if r["avg_val"] is not None else None for r in rows],
        }
        
        # Add corner columns (c1 to c4 based on your data)
        for col in ['c1', 'c2', 'c3', 'c4']:
            if col in rows[0]:
                result[col] = [float(r[col]) if r[col] is not None else None for r in rows]
        
        print(f"DEBUG: Returning data - elevations: {len(result['elevation'])}, c1: {len(result.get('c1', []))}")
        
        return result
        
    except Exception as e:
        print(f"ERROR in get_run: {str(e)}")
        import traceback
        traceback.print_exc()
        return {"error": str(e), "elevation": [], "average": []}

@app.get("/history/{run_id}/boiler-params")
def get_boiler_params(run_id: int):
    conn = get_db()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM boiler_mill_params WHERE run_id = %s", (run_id,))
    data = cur.fetchone()
    cur.close()
    conn.close()
    return data or {}

@app.get("/history/{run_id}/coal-mill-params")
def get_coal_mill_params(run_id: int):
    conn = get_db()
    cur = conn.cursor(dictionary=True)
    cur.execute("SELECT * FROM coal_mill_params WHERE run_id = %s ORDER BY mill", (run_id,))
    data = cur.fetchall()
    cur.close()
    conn.close()
    return data


# ─── MISC ─────────────────────────────────────────────────────
@app.get("/compare")
def compare_runs(ids: str):
    conn = get_db()
    cur = conn.cursor()
    cur.callproc("sp_get_comparison_data", (ids,))
    data = cur.fetchall()
    conn.close()
    return data


@app.get("/stations")
def get_stations():
    conn = get_db()
    cur = conn.cursor()
    cur.callproc("sp_get_stations")
    data = cur.fetchall()
    conn.close()
    return data

@app.get("/download-template")
def download_template(location: str = Query(...), unit: int = Query(...)):
    import copy
    from fastapi.responses import StreamingResponse
    from openpyxl import load_workbook
    import io

    template_path = "templates/FTM_template.xlsx"   # ← adjust to your actual template path

    # Load with keep_vba=False, data_only=False — preserves ALL formatting and formulas
    wb = load_workbook(template_path)
    ws = wb.active

    # Save existing styles from A2 and B2 before writing
    # (openpyxl preserves style when you only set .value)
    ws["A2"].value = location
    ws["B2"].value = f"Unit-{unit}"

    # Stream the file back without saving to disk
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"FTM_template_{location}_Unit{unit}.xlsx"
    headers = {
        "Content-Disposition": f'attachment; filename="{filename}"'
    }
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/units/{station_id}")
def get_units(station_id: int):
    conn = get_db()
    cur = conn.cursor()
    cur.callproc("sp_get_units_by_station", (station_id,))
    data = cur.fetchall()
    conn.close()
    return data


@app.get("/mapping-dates")
def get_mapping_dates(station: str = Query(...), unit: int = Query(...)):
    conn = get_db()
    cur = conn.cursor(dictionary=True)
    cur.callproc("sp_get_mapping_dates", (station, unit))
    
    rows = []
    for result_set in cur.stored_results():
        rows = result_set.fetchall()
    
    cur.close()
    conn.close()
    
    # Convert dates to string format explicitly
    formatted_rows = []
    for row in rows:
        formatted_row = {}
        for key, value in row.items():
            if key == 'run_date' and value is not None:
                # Convert date to YYYY-MM-DD string
                if hasattr(value, 'strftime'):
                    formatted_row[key] = value.strftime('%Y-%m-%d')
                else:
                    formatted_row[key] = str(value)[:10]
            else:
                formatted_row[key] = value
        formatted_rows.append(formatted_row)
    
    print(f"Returning {len(formatted_rows)} dates for {station}, unit {unit}")
    print(f"First few dates: {[r['run_date'] for r in formatted_rows[:5]]}")
    
    return formatted_rows

@app.delete("/runs/{run_id}")
def delete_run(run_id: int):
    try:
        conn = get_db()
        cur = conn.cursor()

        # Delete dependent data first (IMPORTANT for FK safety)
        cur.execute("DELETE FROM profile_points WHERE run_id = %s", (run_id,))
        cur.execute("DELETE FROM boiler_mill_params WHERE run_id = %s", (run_id,))
        cur.execute("DELETE FROM coal_mill_params WHERE run_id = %s", (run_id,))
        
        # Finally delete run
        cur.execute("DELETE FROM runs WHERE run_id = %s", (run_id,))

        conn.commit()
        conn.close()

        return {"message": f"Run {run_id} deleted successfully"}

    except Exception as e:
        return {"error": str(e)}
    
@app.get("/get-upload-log")
def get_upload_log():
    conn = get_db()
    cur = conn.cursor()
    cur.callproc("sp_get_event_log")
    rows = cur.fetchall()
    conn.close()
    return rows


class LoginRequest(BaseModel):
    username: str
    password: str


class ElevationProfileRow(BaseModel):
    elevation_m: float = Field(..., description="Elevation in meters")
    values: Dict[str, Optional[float]] = Field(default_factory=dict, description="Corner label -> temperature")
    average: Optional[float] = None


class ElevationProfile(BaseModel):
    columns: List[str]
    rows: List[ElevationProfileRow]


class RunMeta(BaseModel):
    run_date: str
    run_time: Optional[str] = None
    load_mw: Optional[float] = None
    total_coal_flow_tph: Optional[float] = None
    total_air_flow_tph: Optional[float] = None
    burner_tilt_deg: Optional[float] = None
    coal_mills_in_service_count: Optional[float] = None
    coal_mills_in_service: Optional[str] = None
    oil_guns_in_service_count: Optional[float] = None
    oil_guns_in_service: Optional[str] = None


class CoalMillParam(BaseModel):
    mill: str
    coal_flow_tph: Optional[float] = None
    pa_flow_tph: Optional[float] = None
    mill_dp_mmwc: Optional[float] = None
    mill_outlet_temp: Optional[float] = None
    mill_current_amp: Optional[float] = None


class UploadRunPayload(BaseModel):
    # API-facing: treat station as a location/name string (e.g. "BSL").
    # We do NOT require callers to know / provide numeric station_id.
    station: Optional[str] = None
    # Backward-compatible: if older clients still send station_id, we accept it.
    station_id: Optional[Any] = None
    unit_id: Any
    uploaded_by: Optional[str] = "system"
    notes: Optional[str] = ""
    run_meta: RunMeta
    elevation_profile: ElevationProfile
    boiler_params: Optional[Dict[str, Optional[float]]] = None
    flue_gas_temps: Optional[Dict[str, Optional[float]]] = None
    coal_mill_params: Optional[List[CoalMillParam]] = None


class GenerateExcelPayload(BaseModel):
    """
    Accepts the same shape as UploadRunPayload so your frontend
    can reuse the same data object it already sends to /upload-run.
    Additional top-level fields (station, unit_id, run_meta, etc.) drive
    the header rows; elevation_profile drives the measurement table;
    boiler_params + flue_gas_temps drive the Boiler & Mill Parameters block;
    coal_mill_params drives the Coal Mill Parameters block.
    """
    station: Optional[str] = None
    unit_id: Any = None
    run_meta: RunMeta
    elevation_profile: ElevationProfile
    boiler_params: Optional[Dict[str, Optional[float]]] = None
    flue_gas_temps: Optional[Dict[str, Optional[float]]] = None
    coal_mill_params: Optional[List[CoalMillParam]] = None


class BatchGenerateExcelPayload(BaseModel):
    """Wrapper for multiple run payloads"""
    runs: List[GenerateExcelPayload]


def _v(val) -> Any:
    """Return numeric if possible, else string, else None."""
    if val is None:
        return None
    try:
        f = float(val)
        return int(f) if f == int(f) else f
    except (ValueError, TypeError):
        return str(val) if str(val).strip() else None


def _create_excel_sheet(wb, payload: GenerateExcelPayload, sheet_name: str):
    """Create a single formatted sheet within a workbook"""
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    
    ws = wb.create_sheet(title=sheet_name)
    
    # ── Helpers ───────────────────────────────────────────────
    YELLOW   = "FFFF00"
    CYAN     = "00FFFF"
    DARK_RED = "C00000"
    BLACK    = "000000"
    WHITE    = "FFFFFF"
    GREY     = "D9D9D9"
    BLUE     = "0000FF"
    GREEN    = "00B050"

    def _fill(hex_color: str):
        return PatternFill("solid", fgColor=hex_color)

    def _font(bold=False, color=BLACK, size=10, name="Arial"):
        return Font(bold=bold, color=color, size=size, name=name)

    def _border(style="thin"):
        s = Side(style=style)
        return Border(left=s, right=s, top=s, bottom=s)

    def _center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _left():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _apply(cell, value=None, bold=False, fill=None, font_color=BLACK,
               align="center", border=True, size=10):
        if value is not None:
            cell.value = value
        if fill:
            cell.fill = _fill(fill)
        cell.font = _font(bold=bold, color=font_color, size=size)
        cell.alignment = _center() if align == "center" else _left()
        if border:
            cell.border = _border()

    # ── Column widths (A–P approximately) ────────────────────
    col_widths = {
        "A": 22, "B": 10, "C": 10, "D": 10,
        "E": 10, "F": 10, "G": 10, "H": 10,
        "I": 22, "J": 10, "K": 10, "L": 10,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[7].height = 30

    # ─────────────────────────────────────────────────────────
    # ROW 1 – Title
    # ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:L1")
    c = ws["A1"]
    c.value = "FURNACE FLAME TEMPERATURE MAPPING"
    c.font = _font(bold=True, size=14)
    c.alignment = _center()
    c.fill = _fill(GREY)
    c.border = _border()

    # ─────────────────────────────────────────────────────────
    # ROW 2 – BSL | Unit | Date | Time
    # ─────────────────────────────────────────────────────────
    meta  = payload.run_meta
    loc   = payload.station or ""
    unit  = str(payload.unit_id) if payload.unit_id is not None else ""
    try:
        rd = parse_date_flexible(meta.run_date)
        date_str = rd.strftime("%d.%m.%Y") if rd else meta.run_date
    except Exception:
        date_str = meta.run_date

    _apply(ws["A2"], loc,      bold=True, fill=GREY)
    _apply(ws["B2"], f"Unit-{unit}", bold=True, fill=GREY)
    _apply(ws["C2"], "Date :-",     bold=True, fill=GREY)
    ws.merge_cells("D2:E2")
    _apply(ws["D2"], date_str,      bold=True, fill=YELLOW, font_color="FF0000")
    _apply(ws["F2"], None,  fill=GREY)
    _apply(ws["G2"], None,  fill=GREY)
    _apply(ws["H2"], "Time :-",     bold=True, fill=GREY)
    ws.merge_cells("I2:J2")
    _apply(ws["I2"], meta.run_time or "",  bold=True, fill=YELLOW, font_color="FF0000")

    # ─────────────────────────────────────────────────────────
    # ROW 3 – Load | Coal Flow
    # ─────────────────────────────────────────────────────────
    _apply(ws["A3"], "Load (MW) :-",          bold=True, fill=GREY)
    _apply(ws["B3"], _v(meta.load_mw),        fill=YELLOW)
    _apply(ws["C3"], None, fill=GREY)
    _apply(ws["D3"], "Total Coal flow (TPH) :-", bold=True, fill=GREY)
    ws.merge_cells("E3:F3")
    _apply(ws["E3"], _v(meta.total_coal_flow_tph), fill=YELLOW)
    _apply(ws["G3"], None, fill=GREY)
    _apply(ws["H3"], None, fill=GREY)
    _apply(ws["I3"], _v(meta.total_coal_flow_tph), fill=YELLOW)

    # ─────────────────────────────────────────────────────────
    # ROW 4 – Coal Mills in Service | Air Flow
    # ─────────────────────────────────────────────────────────
    _apply(ws["A4"], "Coal Mills in Service :",   bold=True, fill=GREY)
    _apply(ws["B4"], _v(meta.coal_mills_in_service_count), fill=YELLOW)
    _apply(ws["C4"], meta.coal_mills_in_service or "", fill=YELLOW, font_color=BLUE)
    _apply(ws["D4"], "Total Air flow (TPH) :-",   bold=True, fill=GREY)
    ws.merge_cells("E4:F4")
    _apply(ws["E4"], _v(meta.total_air_flow_tph), fill=YELLOW)
    _apply(ws["G4"], None, fill=GREY)
    _apply(ws["H4"], None, fill=GREY)
    _apply(ws["I4"], _v(meta.total_air_flow_tph), fill=YELLOW)

    # ─────────────────────────────────────────────────────────
    # ROW 5 – Oil Guns | Burner Tilt
    # ─────────────────────────────────────────────────────────
    _apply(ws["A5"], "Oil Guns in Service :",      bold=True, fill=GREY)
    _apply(ws["B5"], _v(meta.oil_guns_in_service_count), fill=YELLOW)
    _apply(ws["C5"], None, fill=GREY)
    _apply(ws["D5"], "Burner Tilt Position (Deg) :-", bold=True, fill=GREY)
    ws.merge_cells("E5:F5")
    _apply(ws["E5"], meta.burner_tilt_deg or "", fill=YELLOW)
    _apply(ws["G5"], None, fill=GREY)
    _apply(ws["H5"], None, fill=GREY)
    _apply(ws["I5"], meta.burner_tilt_deg or "", fill=YELLOW)

    # ─────────────────────────────────────────────────────────
    # ROW 6 – Instrument note
    # ─────────────────────────────────────────────────────────
    ws.merge_cells("A6:L6")
    c = ws["A6"]
    c.value = ("( Flame Temp. in OC - Average ) Instrument Used : Chino Infrared Thermometer"
               " - Model - IR HASNE")
    c.font  = _font(size=9, bold=False)
    c.alignment = _center()
    c.border = _border()

    # ─────────────────────────────────────────────────────────
    # ROW 7 – Column headers
    # ─────────────────────────────────────────────────────────
    ep = payload.elevation_profile
    corner_labels: List[str] = ep.columns

    n_corners = len(corner_labels)
    avg_col_idx   = 2 + n_corners
    elev_r_col    = avg_col_idx + 1

    def _col(one_based: int) -> str:
        return get_column_letter(one_based)

    _apply(ws["A7"], "ELEVATION",  bold=True, fill=GREY)
    for i, lbl in enumerate(corner_labels):
        _apply(ws.cell(row=7, column=2 + i), lbl.upper(), bold=True, fill=GREY)
    _apply(ws.cell(row=7, column=avg_col_idx), "AVERAGE",   bold=True, fill=GREY)
    _apply(ws.cell(row=7, column=elev_r_col),  "ELEVATION", bold=True, fill=GREY)

    # ─────────────────────────────────────────────────────────
    # ROWS 8+ – Profile data
    # ─────────────────────────────────────────────────────────
    data_start_row = 8
    for r_idx, row_data in enumerate(ep.rows):
        xl_row = data_start_row + r_idx
        elev = row_data.elevation_m

        _apply(ws.cell(row=xl_row, column=1), elev, bold=False, fill=GREY)

        vals_numeric = []
        for lbl in corner_labels:
            v = row_data.values.get(lbl)
            if v is not None:
                try:
                    vals_numeric.append(float(v))
                except Exception:
                    vals_numeric.append(None)
            else:
                vals_numeric.append(None)

        valid_vals = [v for v in vals_numeric if v is not None]
        row_max = max(valid_vals) if valid_vals else None
        row_min = min(valid_vals) if valid_vals else None

        for i, lbl in enumerate(corner_labels):
            v = row_data.values.get(lbl)
            cell = ws.cell(row=xl_row, column=2 + i)
            num_v = vals_numeric[i]
            if num_v is not None and row_max is not None and num_v == row_max and row_max != row_min:
                fill_c = YELLOW
            elif num_v is not None and row_min is not None and num_v == row_min and row_max != row_min:
                fill_c = "92D050"
            else:
                fill_c = WHITE
            _apply(cell, _v(v), fill=fill_c)

        avg = row_data.average
        if avg is None and valid_vals:
            avg = sum(valid_vals) / len(valid_vals)
        avg_cell = ws.cell(row=xl_row, column=avg_col_idx)
        _apply(avg_cell, _v(avg), bold=True, fill=CYAN, font_color=BLUE)

        right_lbl = f"{elev} M"
        _apply(ws.cell(row=xl_row, column=elev_r_col), right_lbl, fill=GREY)

    # ─────────────────────────────────────────────────────────
    # NOTE row
    # ─────────────────────────────────────────────────────────
    note_row = data_start_row + len(ep.rows)
    ws.merge_cells(f"A{note_row}:{_col(elev_r_col)}{note_row}")
    c = ws.cell(row=note_row, column=1)
    c.value = "NOTE : - YELLOW  FOR HIGHER TEMPERATURE AND GREEN  FOR LOWER TEMPERATURE"
    c.font  = _font(bold=True, size=9)
    c.alignment = _center()
    c.fill  = _fill(GREY)
    c.border = _border()

    # ─────────────────────────────────────────────────────────
    # BOILER & MILL PARAMETERS block
    # ─────────────────────────────────────────────────────────
    bp_merged = {}
    if payload.boiler_params:
        bp_merged.update(payload.boiler_params)
    if payload.flue_gas_temps:
        bp_merged.update(payload.flue_gas_temps)

    cur_row = note_row + 1

    # Section header
    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    _apply(ws.cell(row=cur_row, column=1), "Boiler & Mill Parameters :", bold=True, fill=GREY)
    _apply(ws.cell(row=cur_row, column=5), "L", bold=True, fill=GREY)
    _apply(ws.cell(row=cur_row, column=6), "R", bold=True, fill=GREY)
    ws.merge_cells(f"G{cur_row}:J{cur_row}")
    _apply(ws.cell(row=cur_row, column=7), "FG Temperature Parameters", bold=True, fill=GREY)
    _apply(ws.cell(row=cur_row, column=11), "L", bold=True, fill=GREY)
    _apply(ws.cell(row=cur_row, column=12), "R", bold=True, fill=GREY)
    cur_row += 1

    def _bp_row(label_left, key_l, key_r, label_right=None, key_r2_l=None, key_r2_r=None):
        nonlocal cur_row
        ws.merge_cells(f"A{cur_row}:D{cur_row}")
        _apply(ws.cell(row=cur_row, column=1), label_left, align="left", fill=GREY)
        _apply(ws.cell(row=cur_row, column=5), _v(bp_merged.get(key_l)) if key_l else None, fill=YELLOW)
        _apply(ws.cell(row=cur_row, column=6), _v(bp_merged.get(key_r)) if key_r else None, fill=YELLOW)
        if label_right:
            ws.merge_cells(f"G{cur_row}:J{cur_row}")
            _apply(ws.cell(row=cur_row, column=7), label_right, align="left", fill=GREY)
            _apply(ws.cell(row=cur_row, column=11), _v(bp_merged.get(key_r2_l)) if key_r2_l else None, fill=YELLOW)
            _apply(ws.cell(row=cur_row, column=12), _v(bp_merged.get(key_r2_r)) if key_r2_r else None, fill=YELLOW)
        cur_row += 1

    _bp_row("Main Steam Pressure (Kg/cm2) :",
            "main_steam_pressure_l", "main_steam_pressure_r",
            "FG Temp after DPSH (L/R ) 0C :",
            "fg_temp_after_dpsh_l", "fg_temp_after_dpsh_r")

    _bp_row("Main Steam Flow (TPH) :",
            "main_steam_flow_l", "main_steam_flow_r",
            "FG Temp after PSH (L/R) 0C :",
            "fg_temp_after_psh_l", "fg_temp_after_psh_r")

    _bp_row("Superheat Spray (TPH) :",
            "superheat_spray_l", "superheat_spray_r",
            "FG Temp after RH (L/R) 0C :",
            "fg_temp_after_rh_l", "fg_temp_after_rh_r")

    _bp_row("Re-heat Spray (TPH) :",
            "reheat_spray_l", "reheat_spray_r",
            "FG Temp after HSH (L/R) 0C :",
            "fg_temp_after_hsh_l", "fg_temp_after_hsh_r")

    _bp_row("O2  at APH Inlet (PCR) % :",
            "o2_aph_inlet_pcr_l", "o2_aph_inlet_pcr_r",
            "FG Temp after Economizer (L/R)  0C :",
            "fg_temp_after_eco_l", "fg_temp_after_eco_r")

    _bp_row("Wind Box DP - A / B (mmwcl) :",
            "wind_box_dp_l", "wind_box_dp_r",
            None, None, None)

    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    _apply(ws.cell(row=cur_row, column=1), "Total PA flow2 (TPH)", align="left", fill=GREY)
    ws.merge_cells(f"E{cur_row}:F{cur_row}")
    _apply(ws.cell(row=cur_row, column=5), _v(bp_merged.get("total_pa_flow")), fill=YELLOW)
    ws.merge_cells(f"G{cur_row}:J{cur_row}")
    _apply(ws.cell(row=cur_row, column=7), "FG Temp after APH (L/R)  0C :", align="left", fill=GREY)
    _apply(ws.cell(row=cur_row, column=11), _v(bp_merged.get("fg_temp_after_aph_l")), fill=YELLOW)
    _apply(ws.cell(row=cur_row, column=12), _v(bp_merged.get("fg_temp_after_aph_r")), fill=YELLOW)
    cur_row += 1

    ws.merge_cells(f"A{cur_row}:D{cur_row}")
    _apply(ws.cell(row=cur_row, column=1), "Total PA flow (TPH)", align="left", fill=GREY)
    ws.merge_cells(f"E{cur_row}:F{cur_row}")
    _apply(ws.cell(row=cur_row, column=5), _v(bp_merged.get("total_pa_flow")), fill=YELLOW)
    cur_row += 1

    # ─────────────────────────────────────────────────────────
    # COAL MILL PARAMETERS block
    # ─────────────────────────────────────────────────────────
    coal_mills = payload.coal_mill_params or []

    ws.merge_cells(f"A{cur_row}:C{cur_row}")
    _apply(ws.cell(row=cur_row, column=1), "Coal Mill Parameters : -", bold=True, fill=GREY)
    cur_row += 1

    if coal_mills:
        _apply(ws.cell(row=cur_row, column=1), "COAL MILLS", bold=True, fill=GREY)
        for i, mill in enumerate(coal_mills):
            _apply(ws.cell(row=cur_row, column=2 + i),
                   f"Coal Mill - {mill.mill}", bold=True, fill=GREY)
        cur_row += 1

        param_rows = [
            ("Coal Flow (TPH)", "coal_flow_tph"),
            ("PA Flow (TPH)",   "pa_flow_tph"),
            ("Mill DP (mmwcl)", "mill_dp_mmwc"),
            ("Mill Outlet Temp (OC)", "mill_outlet_temp"),
            ("Mill Current (Amp)",    "mill_current_amp"),
        ]
        for label, key in param_rows:
            _apply(ws.cell(row=cur_row, column=1), label, align="left", fill=GREY)
            for i, mill in enumerate(coal_mills):
                val = getattr(mill, key, None)
                _apply(ws.cell(row=cur_row, column=2 + i), _v(val), fill=YELLOW)
            cur_row += 1

    # "end" marker
    ws.merge_cells(f"A{cur_row}:C{cur_row}")
    c = ws.cell(row=cur_row, column=1)
    c.value = "end"
    c.font  = _font(bold=True)
    c.alignment = _center()
    c.border = _border()


@app.post("/generate-excel")
def generate_excel(payload: BatchGenerateExcelPayload):
    """
    Build a fully-formatted FTM Excel workbook from JSON data
    and stream it back to the caller as a .xlsx download.
    
    Creates multiple sheets if multiple runs are provided,
    or a single sheet for a single run.
    """
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    import io

    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    
    runs = payload.runs
    
    for idx, run_payload in enumerate(runs):
        meta = run_payload.run_meta
        try:
            rd = parse_date_flexible(meta.run_date)
            date_str = rd.strftime("%d.%m.%Y") if rd else meta.run_date
        except Exception:
            date_str = meta.run_date or f"Run_{idx + 1}"
        
        # Create sheet name (max 31 chars for Excel compatibility)
        sheet_name = date_str.replace("/", "-").replace(".", "-")[:31]
        
        # Handle duplicate sheet names
        if sheet_name in wb.sheetnames:
            sheet_name = f"{sheet_name[:28]}_{idx + 1}"
        
        _create_excel_sheet(wb, run_payload, sheet_name)
    
    # If no sheets were created (empty payload), add a default sheet
    if len(wb.sheetnames) == 0:
        wb.create_sheet("Empty")
    
    # Stream response
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Build filename from first run
    if runs:
        first_meta = runs[0].run_meta
        loc = runs[0].station or ""
        unit = str(runs[0].unit_id) if runs[0].unit_id is not None else ""
        date_safe = (first_meta.run_date or "").replace("/", "-").replace(".", "-")
        filename = f"FTM_{loc}_Unit{unit}_{date_safe}.xlsx"
    else:
        filename = "FTM_export.xlsx"
    
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


def _fetch_stations(conn) -> List[Dict[str, Any]]:
    """
    Robustly fetch stations from `sp_get_stations()`.
    This repo's existing `/stations` endpoint uses `callproc` + `fetchall()`,
    so we mirror that behavior and map tuples to dicts via cursor.description.
    """
    cur = conn.cursor()
    cur.callproc("sp_get_stations")
    rows = cur.fetchall() or []
    cols = [d[0] for d in (cur.description or [])]
    cur.close()
    if not cols:
        return []
    out: List[Dict[str, Any]] = []
    for r in rows:
        try:
            out.append(dict(zip(cols, r)))
        except Exception:
            continue
    return out


def _resolve_station_id_from_runs(conn, location_text: str) -> Optional[int]:
    """
    Fallback: infer station_id by matching existing runs.location.
    This is useful when `sp_get_stations()` is unavailable or returns no rows.
    """
    s = (location_text or "").strip()
    if not s:
        return None
    try:
        cur = conn.cursor(dictionary=True)
        cur.execute(
            """
            SELECT station_id
            FROM runs
            WHERE location = %s
            ORDER BY run_id DESC
            LIMIT 1
            """,
            (s,),
        )
        row = cur.fetchone()
        if row and row.get("station_id") is not None:
            return int(row["station_id"])
        cur.execute(
            """
            SELECT station_id
            FROM runs
            WHERE location LIKE %s
            ORDER BY run_id DESC
            LIMIT 1
            """,
            (f"%{s}%",),
        )
        row = cur.fetchone()
        cur.close()
        if row and row.get("station_id") is not None:
            return int(row["station_id"])
    except Exception:
        return None
    return None


def _resolve_station_id(conn, station_value: Any) -> Optional[int]:
    """
    Accept station id as int / numeric string, or station name.
    Returns an integer station_id if resolvable, else None.
    """
    if station_value is None:
        return None
    if isinstance(station_value, int):
        return station_value
    s = str(station_value).strip()
    if not s:
        return None
    if re.fullmatch(r"\d+", s):
        try:
            return int(s)
        except Exception:
            pass

    # Name path: fetch stations and match by common fields (supports abbreviations)
    try:
        rows = _fetch_stations(conn)
        s_l = re.sub(r"\s+", " ", s).lower()

        def _cand_station_id(r: dict) -> Optional[int]:
            cand = r.get("station_id") or r.get("id") or r.get("stationId")
            if cand is None:
                return None
            if isinstance(cand, int):
                return cand
            cs = str(cand).strip()
            if re.fullmatch(r"\d+", cs):
                return int(cs)
            return None

        # Prefer exact matches first
        for r in rows:
            for k in ("name", "station", "station_name", "location"):
                val = (r.get(k) or "")
                name = re.sub(r"\s+", " ", str(val)).strip()
                if name and name.lower() == s_l:
                    sid = _cand_station_id(r)
                    if sid is not None:
                        return sid

        # Then allow abbreviation / partial matches (e.g. "BSL" matches "BSL Unit-5")
        for r in rows:
            hay = []
            for k in ("name", "station", "station_name", "location"):
                v = (r.get(k) or "")
                vv = re.sub(r"\s+", " ", str(v)).strip()
                if vv:
                    hay.append(vv.lower())
            if not hay:
                continue
            if any(h.startswith(s_l) or s_l in h for h in hay):
                sid = _cand_station_id(r)
                if sid is not None:
                    return sid
    except Exception:
        # If station resolution fails, caller will return a helpful error
        return _resolve_station_id_from_runs(conn, s)

    # If station master returned nothing, try inferring from existing runs
    if not rows:
        return _resolve_station_id_from_runs(conn, s)

    return None


def _resolve_station_location(conn, station_value: Any) -> str:
    """
    Best-effort location string for `runs.location`.
    - If UI sent a non-numeric station name, use it.
    - If UI sent an id, attempt to look up the station name via sp_get_stations.
    """
    if station_value is None:
        return ""
    s = str(station_value).strip()
    if not s:
        return ""
    if not re.fullmatch(r"\d+", s):
        return s

    # Numeric id: try lookup to a name/location field
    try:
        rows = _fetch_stations(conn)
        station_id_int = int(s)
        for r in rows:
            cand = r.get("station_id") or r.get("id") or r.get("stationId")
            try:
                cand_int = int(cand)
            except Exception:
                continue
            if cand_int == station_id_int:
                return (r.get("location") or r.get("name") or r.get("station_name") or r.get("station") or "").strip()
    except Exception:
        return ""

    return ""


@app.post("/upload-run")
def upload_run_json(payload: UploadRunPayload):
    """
    JSON ingest endpoint used by the "Create Run" UI (no Excel file upload).
    Normalizes run_date to YYYY-MM-DD and persists run + profile + params.
    """
    try:
        run_date_obj = parse_date_flexible(payload.run_meta.run_date)
        if not run_date_obj:
            return {"error": "Invalid run_date"}

        try:
            unit_id_int = int(payload.unit_id)
        except Exception:
            return {"error": "Invalid unit_id"}

        # Convert elevation rows to the same shape expected by sp_add_run_points_bulk.
        # Use c1/c2/... keys derived from "Corner <n>" labels (or any label containing a number).
        points: List[Dict[str, Any]] = []
        for r in payload.elevation_profile.rows:
            elev = r.elevation_m
            point: Dict[str, Any] = {"elevation": str(elev)}
            # values is { "Corner 1": 1000, ... }
            for label, temp in (r.values or {}).items():
                if temp is None:
                    continue
                m = re.search(r"(\d+)", str(label))
                if m:
                    # match excel upload shape (stringy numbers are ok for MySQL JSON parsing)
                    point[f"c{m.group(1)}"] = float(temp)
            avg = r.average
            if avg is None:
                present = [v for k, v in point.items() if k != "elevation" and isinstance(v, (int, float))]
                avg = (sum(present) / len(present)) if present else None
            point["avg"] = float(avg) if avg is not None else None
            point["avg_val"] = float(avg) if avg is not None else None  # tolerate either naming convention
            points.append(point)

        conn = get_db()
        cur = conn.cursor(dictionary=True)

        # Station concept for API: location/name string.
        location_text = (payload.station or "").strip()
        if not location_text and payload.station_id is not None:
            # If older clients sent station_id as a name, preserve it as location.
            location_text = str(payload.station_id).strip()

        # DB still requires station_id INT. We infer it if possible, otherwise default.
        station_id_int = None
        if payload.station_id is not None:
            station_id_int = _resolve_station_id(conn, payload.station_id)
        if station_id_int is None and location_text:
            station_id_int = _resolve_station_id_from_runs(conn, location_text)
        if station_id_int is None:
            # Last resort: use a stable default. Location + unit + date still disambiguate runs.
            station_id_int = 1

        # Create run
        cur.callproc(
            "sp_create_run",
            (
                station_id_int,
                unit_id_int,
                "manual_entry",
                datetime.now(),
                run_date_obj,
                payload.uploaded_by or "system",
                payload.notes or "",
                location_text,
                unit_id_int,  # `runs.unit` is INT in your SP
            ),
        )

        run_id = None
        for result_set in cur.stored_results():
            row = result_set.fetchone()
            if row:
                run_id = row.get("run_id")
                break

        if run_id is None:
            cur.close()
            conn.close()
            return {"error": "Failed to create run"}

        # Persist profile points
        cur.callproc(
            "sp_add_run_points_bulk",
            (run_id, location_text, unit_id_int, json.dumps(points)),
        )

        # Persist boiler + flue gas params (same table in this backend)
        merged_params: Dict[str, Any] = {}
        if payload.boiler_params:
            merged_params.update(payload.boiler_params)
        if payload.flue_gas_temps:
            merged_params.update(payload.flue_gas_temps)
        if merged_params:
            upsert_boiler_mill_params(cur, run_id, merged_params)

        # Persist coal mills
        if payload.coal_mill_params:
            upsert_coal_mill_params(cur, run_id, [m.model_dump() for m in payload.coal_mill_params])

        conn.commit()
        cur.close()
        conn.close()

        return {"message": "Run uploaded successfully", "run_id": run_id}

    except Exception as exc:
        import traceback
        return {"error": str(exc), "trace": traceback.format_exc()}


@app.post("/login")
def login(data: LoginRequest):
    try:
        conn = get_db()
        cur = conn.cursor()

        # Direct query for testing
        cur.execute("SELECT * FROM user_master WHERE username = %s AND password = %s", 
                   (data.username, data.password))

        rows = cur.fetchall()
        
        # Get column names
        columns = [desc[0] for desc in cur.description] if cur.description else []
        
        conn.close()

        if len(rows) > 0:
            row_dict = dict(zip(columns, rows[0]))
            location = row_dict.get("location")
            
            return {
                "message": "Login successful",
                "username": data.username,
                "location": location,
                "role": row_dict.get("role")
            }
        else:
            return {
                "message": "Login failed",
                "username": data.username,
                "location": None,
                "role": None
            }

    except Exception as e:
        print(f"Error: {str(e)}")
        return {"error": str(e)}